from fastapi import FastAPI, HTTPException, BackgroundTasks, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, Response
from pydantic import BaseModel, Field
from typing import Optional, List
import os
import io
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import backend.database as db
import backend.scraper as scraper
import backend.predictor as predictor
import backend.alerts as alerts
import backend.scheduler as scheduler

VERCEL = os.environ.get("VERCEL", "0") == "1"

app = FastAPI(title="E-Commerce Price Tracker API", version="1.0.0")

# Enable CORS for local development
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Startup & Shutdown events
@app.on_event("startup")
def startup_event():
    db.init_db()
    # Pre-populate sample products if empty, to make the app gorgeous from the start
    products = db.get_products()
    # Also detect leftover test data with the generic mock title and reset
    has_only_test_data = (
        len(products) == 1 and
        products[0]["title"] == "Premium Smart Gadget 5G (Midnight Black)"
    )
    if not products or has_only_test_data:
        if has_only_test_data:
            print("Clearing leftover test data, pre-populating real sample products...")
            db.clear_mock_data()
            db.init_db()
        else:
            print("Pre-populating database with sample products for instant wow factor...")
        samples = [
            ("https://www.amazon.in/dp/B0CHX1W1XY", "Apple iPhone 15 Pro (128 GB) - Blue Titanium", 129900.0, 120000.0, "https://images.unsplash.com/photo-1695048133142-1a20484d2569?auto=format&fit=crop&w=600&q=80", 4.7, "Amazon"),
            ("https://www.flipkart.com/samsung-galaxy-s24-ultra-5g/p/itmd718fb4839cf9", "Samsung Galaxy S24 Ultra 5G (Titanium Gray, 256GB)", 119999.0, 115000.0, "https://images.unsplash.com/photo-1610945265064-0e34e5519bbf?auto=format&fit=crop&w=600&q=80", 4.6, "Flipkart"),
            ("https://www.amazon.in/dp/B09TKBQR7B", "Sony WH-1000XM5 Wireless Noise Cancelling Headphones", 29990.0, 26000.0, "https://images.unsplash.com/photo-1546435770-a3e426bf472b?auto=format&fit=crop&w=600&q=80", 4.5, "Amazon")
        ]
        for url, title, price, target, img, rating, store in samples:
            p_id = db.add_product(title, url, store, img, rating, price, target)
            if p_id:
                # Pre-populate 45 days of price history
                history = scraper.generate_price_history(price, days=45)
                for entry in history:
                    db.add_price_history(p_id, entry["price"], entry["timestamp"])
                # Add sample alert
                db.add_alert(p_id, "user@example.com", target)

    # Start scheduler if not running on Vercel
    if not VERCEL:
        scheduler.start_scheduler()
    else:
        print("[Scheduler] Bypassing background thread scheduler on Vercel.")

@app.on_event("shutdown")
def shutdown_event():
    if not VERCEL:
        scheduler.stop_scheduler()

# Pydantic models
class ProductCreate(BaseModel):
    url: str
    target_price: Optional[float] = None
    email: Optional[str] = None
    force_mock: bool = False

class AlertCreate(BaseModel):
    product_id: int
    email: str
    target_price: float

class SMTPConfig(BaseModel):
    smtp_server: str
    smtp_port: int
    sender_email: str
    sender_password: str
    enabled: bool

# Curated product catalog for one-click tracking
PRODUCT_CATALOG = [
    # Smartphones
    {"id": "cat_1", "category": "Smartphones", "title": "Apple iPhone 15 Pro (128 GB) - Blue Titanium", "price": 129900.0, "url": "https://www.amazon.in/dp/B0CHX1W1XY", "store": "Amazon", "rating": 4.7, "image_url": "https://images.unsplash.com/photo-1695048133142-1a20484d2569?auto=format&fit=crop&w=600&q=80"},
    {"id": "cat_2", "category": "Smartphones", "title": "Samsung Galaxy S24 Ultra 5G (256GB, Titanium Gray)", "price": 119999.0, "url": "https://www.flipkart.com/samsung-galaxy-s24-ultra-5g/p/itmd718fb4839cf9", "store": "Flipkart", "rating": 4.6, "image_url": "https://images.unsplash.com/photo-1610945265064-0e34e5519bbf?auto=format&fit=crop&w=600&q=80"},
    {"id": "cat_3", "category": "Smartphones", "title": "OnePlus 12 5G (Flowy Emerald, 12GB RAM, 256GB)", "price": 64999.0, "url": "https://www.amazon.in/dp/B0CQYQF9WJ", "store": "Amazon", "rating": 4.5, "image_url": "https://images.unsplash.com/photo-1598300042247-d088f8ab3a91?auto=format&fit=crop&w=600&q=80"},
    {"id": "cat_4", "category": "Smartphones", "title": "Google Pixel 8 Pro (Obsidian, 128GB)", "price": 89999.0, "url": "https://www.flipkart.com/google-pixel-8-pro/p/itmb8c7c0c59b1f2", "store": "Flipkart", "rating": 4.4, "image_url": "https://images.unsplash.com/photo-1598300042247-d088f8ab3a91?auto=format&fit=crop&w=600&q=80"},
    # Laptops
    {"id": "cat_5", "category": "Laptops", "title": "Apple MacBook Air 13\" M3 Chip (8GB RAM, 256GB SSD)", "price": 114900.0, "url": "https://www.amazon.in/dp/B0CTGT4DFW", "store": "Amazon", "rating": 4.8, "image_url": "https://images.unsplash.com/photo-1517336714731-489689fd1ca8?auto=format&fit=crop&w=600&q=80"},
    {"id": "cat_6", "category": "Laptops", "title": "Dell XPS 15 Intel Core i7 13th Gen (16GB, 512GB SSD)", "price": 149990.0, "url": "https://www.amazon.in/dp/B0BXBHBQJ6", "store": "Amazon", "rating": 4.6, "image_url": "https://images.unsplash.com/photo-1593642632559-0c6d3fc62b89?auto=format&fit=crop&w=600&q=80"},
    {"id": "cat_7", "category": "Laptops", "title": "HP Pavilion x360 14\" Touch Intel i5 (8GB, 512GB)", "price": 69990.0, "url": "https://www.flipkart.com/hp-pavilion-x360/p/itm7ab4f42e09f0e", "store": "Flipkart", "rating": 4.3, "image_url": "https://images.unsplash.com/photo-1587614382346-4ec70e388b28?auto=format&fit=crop&w=600&q=80"},
    # Audio
    {"id": "cat_8", "category": "Audio", "title": "Sony WH-1000XM5 Wireless Noise Cancelling Headphones", "price": 29990.0, "url": "https://www.amazon.in/dp/B09TKBQR7B", "store": "Amazon", "rating": 4.5, "image_url": "https://images.unsplash.com/photo-1546435770-a3e426bf472b?auto=format&fit=crop&w=600&q=80"},
    {"id": "cat_9", "category": "Audio", "title": "JBL Tune 510BT Wireless On-Ear Headphones (Black)", "price": 3499.0, "url": "https://www.amazon.in/dp/B097QR8CMJ", "store": "Amazon", "rating": 4.3, "image_url": "https://images.unsplash.com/photo-1484704849700-f032a568e944?auto=format&fit=crop&w=600&q=80"},
    {"id": "cat_10", "category": "Audio", "title": "Apple AirPods Pro 2nd Generation (USB-C)", "price": 24900.0, "url": "https://www.amazon.in/dp/B0BSHF7WHJ", "store": "Amazon", "rating": 4.7, "image_url": "https://images.unsplash.com/photo-1588423771073-b8903febb85b?auto=format&fit=crop&w=600&q=80"},
    # Smartwatches
    {"id": "cat_11", "category": "Smartwatches", "title": "Apple Watch Series 9 (GPS, 41mm, Midnight Aluminium)", "price": 41900.0, "url": "https://www.amazon.in/dp/B0CHX3QBCH", "store": "Amazon", "rating": 4.7, "image_url": "https://images.unsplash.com/photo-1434493789847-2f02dc6ca35d?auto=format&fit=crop&w=600&q=80"},
    {"id": "cat_12", "category": "Smartwatches", "title": "Samsung Galaxy Watch 6 Classic 43mm (Black)", "price": 36999.0, "url": "https://www.flipkart.com/samsung-galaxy-watch-6-classic/p/itmc61b4d2f9a2d9", "store": "Flipkart", "rating": 4.5, "image_url": "https://images.unsplash.com/photo-1523275335684-37898b6baf30?auto=format&fit=crop&w=600&q=80"},
    # Tablets
    {"id": "cat_13", "category": "Tablets", "title": "Apple iPad Air 11\" M2 Chip (128GB, Wi-Fi, Blue)", "price": 74900.0, "url": "https://www.amazon.in/dp/B0D3J7DK3N", "store": "Amazon", "rating": 4.8, "image_url": "https://images.unsplash.com/photo-1544244015-0df4b3ffc6b0?auto=format&fit=crop&w=600&q=80"},
    {"id": "cat_14", "category": "Tablets", "title": "Samsung Galaxy Tab S9 FE (256GB, 8GB RAM, Gray)", "price": 49999.0, "url": "https://www.flipkart.com/samsung-galaxy-tab-s9-fe/p/itmb8f97da7d45d4", "store": "Flipkart", "rating": 4.4, "image_url": "https://images.unsplash.com/photo-1561154464-82e9adf32764?auto=format&fit=crop&w=600&q=80"},
    {"id": "cat_15", "category": "Tablets", "title": "All-new Kindle Paperwhite (16GB) - 6.8\" Display", "price": 14999.0, "url": "https://www.amazon.in/dp/B09TMF6742", "store": "Amazon", "rating": 4.5, "image_url": "https://images.unsplash.com/photo-1544716278-ca5e3f4abd8c?auto=format&fit=crop&w=600&q=80"},
    # Cameras
    {"id": "cat_16", "category": "Cameras", "title": "Sony Alpha ZV-E10 Mirrorless Camera (Body Only)", "price": 54990.0, "url": "https://www.amazon.in/dp/B09BFPGC7W", "store": "Amazon", "rating": 4.5, "image_url": "https://images.unsplash.com/photo-1516035069371-29a1b244cc32?auto=format&fit=crop&w=600&q=80"},
    {"id": "cat_17", "category": "Cameras", "title": "Canon EOS R50 Mirrorless Camera (Black, Body Only)", "price": 62990.0, "url": "https://www.amazon.in/dp/B0BW1JZ5WB", "store": "Amazon", "rating": 4.6, "image_url": "https://images.unsplash.com/photo-1502920917128-1aa500764cbd?auto=format&fit=crop&w=600&q=80"},
    {"id": "cat_18", "category": "Cameras", "title": "GoPro HERO12 Black Action Camera (5.3K)", "price": 34990.0, "url": "https://www.amazon.in/dp/B0CF8PFHX5", "store": "Amazon", "rating": 4.4, "image_url": "https://images.unsplash.com/photo-1564466809058-bf4114d55352?auto=format&fit=crop&w=600&q=80"},
]

@app.get("/api/catalog")
def get_catalog():
    """Returns the curated product catalog for one-click tracking."""
    return PRODUCT_CATALOG

# API Routes
@app.get("/api/products")
def get_products():
    """Returns lightweight product list instantly — NO ML computation here."""
    products = db.get_products()
    results = []
    for p in products:
        history_count = db.get_price_history_count(p["id"])
        first_price = db.get_first_price(p["id"])
        initial_price = first_price if first_price else p["current_price"]
        price_drop = initial_price - p["current_price"]
        drop_pct = round((price_drop / initial_price) * 100, 1) if initial_price > 0 else 0

        results.append({
            **p,
            "price_history_count": history_count,
            "drop_percentage": drop_pct,
            # Placeholder stats — frontend fetches real stats via /stats endpoint lazily
            "stats": {
                "min_price": p["current_price"],
                "max_price": p["current_price"],
                "avg_price": p["current_price"],
                "volatility": 0.0,
                "price_drop_prob": 0
            },
            "forecast": []
        })
    return results

@app.get("/api/products/{product_id}/stats")
def get_product_stats(product_id: int):
    """Returns ML stats and forecast for a single product — called lazily by the frontend."""
    p = db.get_product(product_id)
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    history = db.get_price_history(product_id)
    pred_stats = predictor.predict_price_trends(history)

    initial_price = history[0]["price"] if history else p["current_price"]
    price_drop = initial_price - p["current_price"]
    drop_pct = round((price_drop / initial_price) * 100, 1) if initial_price > 0 else 0

    return {
        "drop_percentage": drop_pct,
        "stats": {
            "min_price": pred_stats["min_price"],
            "max_price": pred_stats["max_price"],
            "avg_price": pred_stats["avg_price"],
            "volatility": pred_stats["volatility"],
            "price_drop_prob": pred_stats["price_drop_prob"]
        },
        "forecast": pred_stats["forecast"]
    }

@app.post("/api/products")
def create_product(product_data: ProductCreate, background_tasks: BackgroundTasks):
    url = product_data.url.strip()
    if not url:
        raise HTTPException(status_code=400, detail="URL cannot be empty")
        
    # Attempt scrape
    scraped = scraper.scrape_product(url, force_mock=product_data.force_mock)
    if not scraped:
        raise HTTPException(status_code=400, detail="Failed to scrape details from this URL")
        
    p_id = db.add_product(
        title=scraped["title"],
        url=url,
        store=scraped["store"],
        image_url=scraped["image_url"],
        rating=scraped["rating"],
        current_price=scraped["price"],
        target_price=product_data.target_price
    )
    
    if not p_id:
        raise HTTPException(status_code=400, detail="Product could not be added (maybe duplicate URL)")
        
    # Pre-populate history (30 days) to make ML predictions work instantly
    history = scraper.generate_price_history(scraped["price"], days=30)
    for entry in history:
        # Ignore last entry to keep current scraped price as the latest point
        if entry == history[-1]:
            continue
        db.add_price_history(p_id, entry["price"], entry["timestamp"])
        
    # Add alert if email and target price are provided
    if product_data.email and product_data.target_price:
        db.add_alert(p_id, product_data.email, product_data.target_price)
        
    # Trigger alert check in background
    background_tasks.add_task(alerts.check_and_trigger_alerts)
    
    return {"message": "Product added successfully", "product_id": p_id, "details": scraped}

@app.delete("/api/products/{product_id}")
def delete_product(product_id: int):
    p = db.get_product(product_id)
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    db.delete_product(product_id)
    return {"message": "Product deleted successfully"}

@app.get("/api/products/{product_id}/history")
def get_product_history(product_id: int):
    p = db.get_product(product_id)
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    return db.get_price_history(product_id)

@app.get("/api/products/{product_id}/predict")
def get_product_prediction(product_id: int):
    p = db.get_product(product_id)
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    history = db.get_price_history(product_id)
    return predictor.predict_price_trends(history)

@app.post("/api/products/{product_id}/scrape")
def trigger_scrape(product_id: int, background_tasks: BackgroundTasks):
    p = db.get_product(product_id)
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
        
    scraped = scraper.scrape_product(p["url"], current_price=p["current_price"])
    if not scraped:
        raise HTTPException(status_code=400, detail="Failed to scrape product")
        
    db.update_product_price(product_id, scraped["price"])
    background_tasks.add_task(alerts.check_and_trigger_alerts)
    
    return {"message": "Scrape completed successfully", "price": scraped["price"]}

@app.post("/api/scrape-all")
def scrape_all(background_tasks: BackgroundTasks):
    background_tasks.add_task(scheduler.run_scrape_cycle)
    return {"message": "Background scrape job triggered"}

@app.get("/api/alerts")
def get_alerts():
    return db.get_alerts(only_active=False)

@app.post("/api/alerts")
def create_alert(alert_data: AlertCreate, background_tasks: BackgroundTasks):
    p = db.get_product(alert_data.product_id)
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
        
    alert_id = db.add_alert(alert_data.product_id, alert_data.email, alert_data.target_price)
    
    # Update product target price in product details
    conn = db.get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE products SET target_price = ? WHERE id = ?", (alert_data.target_price, alert_data.product_id))
    conn.commit()
    conn.close()
    
    background_tasks.add_task(alerts.check_and_trigger_alerts)
    return {"message": "Alert added successfully", "alert_id": alert_id}

@app.get("/api/config")
def get_config():
    config = alerts.get_email_config()
    # Mask password for security
    if config.get("sender_password"):
        config["sender_password"] = "********"
    return config

@app.post("/api/config")
def update_config(config_data: SMTPConfig):
    # If password is masked, reload it from current config
    password = config_data.sender_password
    if password == "********":
        current = alerts.get_email_config()
        password = current.get("sender_password", "")
        
    saved = alerts.save_email_config(
        smtp_server=config_data.smtp_server,
        smtp_port=config_data.smtp_port,
        sender_email=config_data.sender_email,
        sender_password=password,
        enabled=config_data.enabled
    )
    # Mask password for return
    saved["sender_password"] = "********"
    return {"message": "SMTP Configuration saved", "config": saved}

@app.get("/api/alerts-preview")
def get_alerts_preview():
    if VERCEL:
        html_preview_path = "/tmp/logs/alerts_preview.html"
    else:
        html_preview_path = os.path.join(os.path.dirname(__file__), "..", "logs", "alerts_preview.html")
        
    if not os.path.exists(html_preview_path):
        return {"html": "<p style='color:#94a3b8; font-family:sans-serif; text-align:center;'>No alerts triggered yet. Add a product, set an alert below its price, and trigger a manual update to see notifications here!</p>"}
    try:
        with open(html_preview_path, "r", encoding="utf-8") as f:
            return {"html": f.read()}
    except Exception as e:
        return {"html": f"<p style='color:red;'>Error loading preview: {str(e)}</p>"}

@app.post("/api/clear-logs")
def clear_logs():
    if VERCEL:
        html_preview_path = "/tmp/logs/alerts_preview.html"
        txt_log_path = "/tmp/logs/alerts.log"
    else:
        html_preview_path = os.path.join(os.path.dirname(__file__), "..", "logs", "alerts_preview.html")
        txt_log_path = os.path.join(os.path.dirname(__file__), "..", "logs", "alerts.log")
    
    for path in [html_preview_path, txt_log_path]:
        if os.path.exists(path):
            try:
                os.remove(path)
            except Exception:
                pass
    return {"message": "Logs cleared"}

@app.post("/api/clear-db")
def clear_database():
    db.clear_mock_data()
    db.init_db()
    return {"message": "Database reset completed"}

@app.get("/api/export")
def export_excel():
    """
    Generates a professionally formatted Excel spreadsheet containing
    all tracked products, current prices, alert conditions, and historical charts.
    """
    products = db.get_products()
    
    # Create Excel Workbook
    wb = openpyxl.Workbook()
    
    # Sheet 1: Products Summary
    ws_summary = wb.active
    ws_summary.title = "Products Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Colors and styles
    purple_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")  # Indigo
    white_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Segoe UI", size=11, bold=True)
    regular_font = Font(name="Segoe UI", size=11)
    title_font = Font(name="Segoe UI", size=16, bold=True, color="1F2937")
    gray_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    border_thin = Side(border_style="thin", color="E5E7EB")
    border_double = Side(border_style="double", color="374151")
    grid_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    bottom_double_border = Border(bottom=border_double, top=border_thin)
    
    # Title Block
    ws_summary["A1"] = "E-Commerce Price Tracker Report"
    ws_summary["A1"].font = title_font
    ws_summary["A2"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary["A2"].font = Font(name="Segoe UI", size=10, italic=True, color="6B7280")
    
    # Headers
    headers = ["ID", "Store", "Product Title", "Current Price (INR)", "Target Price (INR)", "Rating", "Created Date", "Tracked Days"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=4, column=col_num)
        cell.value = header
        cell.fill = purple_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = grid_border
    
    ws_summary.row_dimensions[4].height = 25
    
    row_idx = 5
    for p in products:
        history = db.get_price_history(p["id"])
        tracked_days = len(history)
        
        ws_summary.cell(row=row_idx, column=1, value=p["id"]).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=row_idx, column=2, value=p["store"]).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=row_idx, column=3, value=p["title"])
        
        c_price = ws_summary.cell(row=row_idx, column=4, value=p["current_price"])
        c_price.number_format = "₹#,##0.00"
        c_price.alignment = Alignment(horizontal="right")
        
        t_price = ws_summary.cell(row=row_idx, column=5, value=p["target_price"])
        if p["target_price"]:
            t_price.number_format = "₹#,##0.00"
        else:
            t_price.value = "N/A"
        t_price.alignment = Alignment(horizontal="right")
        
        ws_summary.cell(row=row_idx, column=6, value=p["rating"]).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=row_idx, column=7, value=p["created_at"].split()[0]).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=row_idx, column=8, value=tracked_days).alignment = Alignment(horizontal="center")
        
        # Style rows
        for col_num in range(1, 9):
            c = ws_summary.cell(row=row_idx, column=col_num)
            c.font = regular_font
            c.border = grid_border
            if row_idx % 2 == 0:
                c.fill = gray_fill
                
        row_idx += 1
        
    # Auto-adjust column widths
    for col in ws_summary.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            # Avoid using title row to size column A
            if cell.row < 4:
                continue
            if cell.value:
                # Truncate length of extremely long product titles for spacing
                val_str = str(cell.value)
                if len(val_str) > 40 and cell.column == 3:
                    max_len = max(max_len, 40)
                else:
                    max_len = max(max_len, len(val_str))
        ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # Sheet 2: Historical Raw Data
    ws_history = wb.create_sheet(title="Price Histories")
    ws_history.views.sheetView[0].showGridLines = True
    
    ws_history["A1"] = "Detailed Price Change Records"
    ws_history["A1"].font = title_font
    
    history_headers = ["Product ID", "Product Title", "Recorded Price (INR)", "Timestamp"]
    for col_num, header in enumerate(history_headers, 1):
        cell = ws_history.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")  # Slate Gray
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = grid_border
        
    hist_row_idx = 4
    for p in products:
        history = db.get_price_history(p["id"])
        for h in history:
            ws_history.cell(row=hist_row_idx, column=1, value=p["id"]).alignment = Alignment(horizontal="center")
            ws_history.cell(row=hist_row_idx, column=2, value=p["title"])
            
            price_cell = ws_history.cell(row=hist_row_idx, column=3, value=h["price"])
            price_cell.number_format = "₹#,##0.00"
            price_cell.alignment = Alignment(horizontal="right")
            
            ws_history.cell(row=hist_row_idx, column=4, value=h["timestamp"]).alignment = Alignment(horizontal="center")
            
            for col_num in range(1, 5):
                c = ws_history.cell(row=hist_row_idx, column=col_num)
                c.font = regular_font
                c.border = grid_border
                
            hist_row_idx += 1
            
    for col in ws_history.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 3:
                continue
            if cell.value:
                val_str = str(cell.value)
                if len(val_str) > 40 and cell.column == 2:
                    max_len = max(max_len, 40)
                else:
                    max_len = max(max_len, len(val_str))
        ws_history.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Save to a dynamic bytes output buffer
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    filename = f"ECommerce_Price_Tracker_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return Response(
        stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )

# Mount the static web UI as standard SPA.
# This MUST be mounted at the end so it doesn't hijack API routes
app.mount("/", StaticFiles(directory=os.path.join(os.path.dirname(__file__), "..", "frontend"), html=True), name="frontend")
