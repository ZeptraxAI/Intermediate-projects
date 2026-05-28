import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from app.config import DATA_DIR

LOGS_CSV = DATA_DIR / "logs.csv"
REPORT_EXCEL = DATA_DIR / "health_report.xlsx"

# Columns for our telemetry logs
LOG_COLUMNS = [
    "timestamp",
    "site_id",
    "site_name",
    "url",
    "status_code",
    "response_time_ms",
    "is_up",
    "error_message",
    "check_type",
    "screenshot_path"
]

def append_log(
    site_id: str,
    site_name: str,
    url: str,
    status_code: int,
    response_time_ms: float,
    is_up: bool,
    error_message: str = "",
    check_type: str = "request",
    screenshot_path: str = ""
):
    """
    Appends a new log entry to the CSV log file using Pandas.
    """
    log_entry = {
        "timestamp": datetime.now().isoformat(),
        "site_id": site_id,
        "site_name": site_name,
        "url": url,
        "status_code": status_code,
        "response_time_ms": round(response_time_ms, 2),
        "is_up": is_up,
        "error_message": error_message,
        "check_type": check_type,
        "screenshot_path": screenshot_path
    }
    
    df = pd.DataFrame([log_entry], columns=LOG_COLUMNS)
    
    # If file doesn't exist, write with header. Otherwise, append without header.
    if not LOGS_CSV.exists():
        df.to_csv(LOGS_CSV, index=False)
    else:
        df.to_csv(LOGS_CSV, mode='a', header=False, index=False)

def get_logs_df() -> pd.DataFrame:
    """
    Loads logs from the CSV file and returns a Pandas DataFrame.
    """
    if not LOGS_CSV.exists():
        return pd.DataFrame(columns=LOG_COLUMNS)
    try:
        # Load and ensure types are correct
        df = pd.read_csv(LOGS_CSV)
        # Ensure correct column ordering
        df = df.reindex(columns=LOG_COLUMNS)
        return df
    except Exception as e:
        print(f"Error loading logs CSV: {e}")
        return pd.DataFrame(columns=LOG_COLUMNS)

def export_excel_report() -> str:
    """
    Exports a styled Excel report containing summary analytics and raw logs.
    Returns the path to the generated Excel file.
    """
    df = get_logs_df()
    if df.empty:
        # Write an empty but valid Excel file if no logs exist
        wb = Workbook()
        ws = wb.active
        ws.title = "No Data"
        ws["A1"] = "No logs recorded yet."
        wb.save(REPORT_EXCEL)
        return str(REPORT_EXCEL)

    # 1. Calculate Summary Stats
    total_checks = len(df)
    uptime_ratio = df["is_up"].mean() * 100 if total_checks > 0 else 100.0
    avg_latency = df["response_time_ms"].mean() if total_checks > 0 else 0.0
    slow_checks = len(df[df["response_time_ms"] > 2000])
    down_checks = len(df[df["is_up"] == False])
    
    # Group by site for brief summary
    site_groups = df.groupby("site_id").agg(
        name=("site_name", "first"),
        url=("url", "first"),
        checks=("is_up", "count"),
        uptime=("is_up", lambda x: x.mean() * 100),
        avg_latency=("response_time_ms", "mean")
    ).reset_index()

    # Create openpyxl Workbook
    wb = Workbook()
    
    # --- Worksheet 1: Dashboard / Executive Summary ---
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.views.sheetView[0].showGridLines = True

    # Color Palette - Sleek Corporate Dark Blue Theme
    navy_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    light_blue_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
    green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    title_font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    section_font = Font(name="Segoe UI", size=12, bold=True, color="1B365D")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Segoe UI", size=11, bold=True)
    regular_font = Font(name="Segoe UI", size=11)
    sm_font = Font(name="Segoe UI", size=9, italic=True)

    border_thin = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    border_double_bottom = Border(
        bottom=Side(style='double', color='1B365D'),
        top=Side(style='thin', color='DDDDDD')
    )

    # Title Banner
    ws_summary.merge_cells("A1:E2")
    ws_summary["A1"] = "WEBSITE HEALTH MONITOR - SERVICE REPORT"
    ws_summary["A1"].font = title_font
    ws_summary["A1"].fill = navy_fill
    ws_summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    # Generated Date
    ws_summary["A3"] = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary["A3"].font = sm_font

    # Overview KPI Cards
    kpis = [
        ("Total Health Checks", total_checks, "A5:B5"),
        ("Global Uptime Rate", f"{uptime_ratio:.2f}%", "C5:D5"),
        ("Avg Response Time", f"{avg_latency:.1f} ms", "E5:F5"),
        ("Total Downtime Events", down_checks, "A7:B7"),
        ("Slow Latency Incidents (>2s)", slow_checks, "C7:D7"),
    ]
    
    # Style KPI boxes
    for title, val, cells in kpis:
        start_cell = cells.split(":")[0]
        # We write title to the start cell, value to the row below it
        col_letter = start_cell[0]
        row_num = int(start_cell[1:])
        
        ws_summary[f"{col_letter}{row_num}"] = title
        ws_summary[f"{col_letter}{row_num}"].font = sm_font
        
        val_cell = f"{col_letter}{row_num+1}"
        ws_summary[val_cell] = val
        ws_summary[val_cell].font = Font(name="Segoe UI", size=14, bold=True, color="1B365D")
        ws_summary[val_cell].fill = light_blue_fill
        ws_summary[val_cell].border = border_thin
        ws_summary[val_cell].alignment = Alignment(horizontal="center")

    # Section: Site Performance Table
    ws_summary["A10"] = "Monitored Targets Summary"
    ws_summary["A10"].font = section_font
    
    headers = ["Site Name", "URL", "Total Checks", "Uptime %", "Avg Latency (ms)"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=11, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="left" if col_num <= 2 else "right")
        cell.border = border_thin

    row_idx = 12
    for _, row in site_groups.iterrows():
        ws_summary.cell(row=row_idx, column=1, value=row["name"]).font = regular_font
        ws_summary.cell(row=row_idx, column=2, value=row["url"]).font = regular_font
        ws_summary.cell(row=row_idx, column=3, value=row["checks"]).font = regular_font
        
        # Uptime
        ut_cell = ws_summary.cell(row=row_idx, column=4, value=round(row["uptime"], 2))
        ut_cell.font = bold_font
        ut_cell.alignment = Alignment(horizontal="right")
        if row["uptime"] >= 99.0:
            ut_cell.fill = green_fill
        elif row["uptime"] >= 95.0:
            ut_cell.fill = yellow_fill
        else:
            ut_cell.fill = red_fill
            
        # Latency
        lat_cell = ws_summary.cell(row=row_idx, column=5, value=round(row["avg_latency"], 1))
        lat_cell.font = regular_font
        lat_cell.alignment = Alignment(horizontal="right")
        if row["avg_latency"] > 2000:
            lat_cell.fill = yellow_fill
            
        # Borders
        for c in range(1, 6):
            ws_summary.cell(row=row_idx, column=c).border = border_thin
            
        row_idx += 1

    # Format widths
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # --- Worksheet 2: Detailed Logs ---
    ws_logs = wb.create_sheet(title="Detailed Logs")
    ws_logs.views.sheetView[0].showGridLines = True
    
    log_headers = [c.replace('_', ' ').title() for c in LOG_COLUMNS]
    for col_num, header in enumerate(log_headers, 1):
        cell = ws_logs.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = navy_fill
        cell.border = border_thin
        
    for r_num, row_data in enumerate(df.values, 2):
        for c_num, val in enumerate(row_data, 1):
            cell = ws_logs.cell(row=r_num, column=c_num)
            
            # Format and set value
            if c_num == 1:  # Timestamp formatting
                try:
                    dt = datetime.fromisoformat(str(val))
                    cell.value = dt.strftime("%Y-%m-%d %H:%M:%S")
                except:
                    cell.value = str(val)
            elif c_num == 5:  # Status code
                cell.value = int(val) if pd.notna(val) else -1
            elif c_num == 6:  # Latency
                cell.value = round(float(val), 2) if pd.notna(val) else 0.0
            elif c_num == 7:  # Is Up
                cell.value = bool(val)
                # Color code Is Up
                if val:
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill
            else:
                cell.value = "" if pd.isna(val) else str(val)
                
            cell.font = regular_font
            cell.border = border_thin
            
            # Alignment
            if c_num in [1, 5, 6, 7]:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")

    # Format widths for raw logs
    for col in ws_logs.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        # Set max column width limit to prevent giant columns for URL/Error message
        ws_logs.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 30)

    # Save to disk
    wb.save(REPORT_EXCEL)
    return str(REPORT_EXCEL)
